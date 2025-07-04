## This code analyses the cleaning performance of different pipelines on resting state EEGs and visualises this in three figures:
#   1. scatter plots of signal preservation against artifact removal for all pipelines together
#   2. paired scatter plots for improvements in data quality per pipeline
#   3. topomaps and spectra for detection of berger effect with bayesian t-test to address the strength
# Pipelines to evaluate can be given as functions with EEG files to test on or as already preprocessed eeg recordings to 
# compare to the originals. The quantative metrics per file will be registered in an Excel file, which will be updated if 
# new files are available. For the Berger effect the power spectra are saved to the assigned folder.The figures will be 
# made based on the complete Excel file also with metrics from earlier quality metrics. 

#%%  
# import required toolboxes 
import os
import re
import mne
import pyprep
import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# for visualisation
import h5py
import baycomp
import matplotlib.pyplot as plt 
import seaborn as sns 
from itertools import product 
from scipy.stats import shapiro, ttest_rel, wilcoxon
from mne.viz import plot_topomap
from mne.time_frequency import read_spectrum
import pingouin as pg

# custom functions
from Functions_quality_assessment import assess_eeg_quality, compute_ser_arr
from Functions_load_data import load_data_with_markers, extract_segments, identify_periods

# functions needed for preprocessing steps (adjustable)
from Functies_concept_preprocessing import base_filters, wavelet_enhanced_ica, apply_MWF #custom functions
from mne_icalabel import label_components

#%%  
# input 

EEG_folder = 'EEG'
EEG_folder = r"z:\31122024_datauitgifte_RobertvandenBerg\EEG"
EVENT_folder = 'Eyetracker'
EVENT_folder = r"Z:\31122024_datauitgifte_RobertvandenBerg\EyeTracker"

EEG_pipeline_results = 'pipeline' #add pipelines like this if analysing based on preprocessed files instead of giving the function (also three other places below)

results_file = 'Quality_results.xlsx'
folder_PSDS = 'PSDS'

montage = mne.channels.make_standard_montage("GSN-HydroCel-128")

pipelines = ['MWF', 'IClabel', 'wICA'] #name pipelines functions to evaluate and give an apply function below

new_metrics_buffer = {method: [] for method in pipelines}
new_metrics_buffer['base_filters'] = []
new_metrics_buffer['preprocessing_pipeline'] = [] #add pipelines like this if analysing based on preprocessed files instead of giving the function (also three other place above and below)

save_interval = 1 #adjustable by preference but now all results will be saved after analysing each file

#%% 
# define required functions

# functions of preprocessing pipeline (add functions here to evaluate them and call them in the apply method function)
def apply_MWF_pipeline(raw):
    # apply multichannel wiener filter preprocessing pipeline
    raw_filtered = base_filters(raw)
    noisy_channels = pyprep.NoisyChannels(raw_filtered)
    noisy_channels.find_all_bads()
    bad = noisy_channels.get_bads()
    bad = list(map(str, bad))

    raw_filtered.info["bads"] += bad

    raw_MWF= apply_MWF(raw_filtered)

    raw_MWF.set_eeg_reference(ref_channels="average", verbose='CRITICAL')
    raw_MWF.interpolate_bads()
    return raw_MWF

def apply_BF(raw):
    # apply standard filters as preprocessing
    raw_filtered = base_filters(raw)
    raw_filtered.set_eeg_reference(ref_channels="average", verbose='CRITICAL')
    return raw_filtered

def apply_IC(raw):
    # apply preprocessing pipeline with IClabel
    raw_filtered = base_filters(raw)
    noisy_channels = pyprep.NoisyChannels(raw_filtered)
    noisy_channels.find_all_bads()
    bad = noisy_channels.get_bads()
    bad = list(map(str, bad))

    raw_filtered.info["bads"] += bad

    ica_raw = raw_filtered.copy().filter(l_freq=1, h_freq=None)
    ica_raw = ica_raw.set_eeg_reference("average")

    ica = mne.preprocessing.ICA(n_components=15, max_iter="auto", method="infomax", random_state=97, fit_params=dict(extended=True))
    ica.fit(ica_raw)

    ic_labels = label_components(ica_raw, ica, method="iclabel")
    labels = ic_labels["labels"]
    exclude_idx = [idx for idx, label in enumerate(labels) if label not in ["brain", "other"]]
    reconst_raw = raw_filtered.copy()
    ica.apply(reconst_raw, exclude = exclude_idx)

    reconst_raw.set_eeg_reference(ref_channels="average", verbose='CRITICAL')
    reconst_raw.interpolate_bads()
    return reconst_raw

def apply_wICA(raw):
    # apply wavelet enhanced ICA preprocessing pipeline
    raw_filtered = base_filters(raw)
    noisy_channels = pyprep.NoisyChannels(raw_filtered)
    noisy_channels.find_all_bads()
    bad = noisy_channels.get_bads()
    bad = list(map(str, bad))

    raw_filtered.info["bads"] += bad

    ica_raw = raw_filtered.copy()

    raw_wICA = wavelet_enhanced_ica(ica_raw)
    raw_wICA.set_eeg_reference(ref_channels="average", verbose='CRITICAL')
    raw_wICA.interpolate_bads()    
    return raw_wICA

def apply_method(data, method): # add additional preprocessings functions here to evaluate if wanted
    # general function to apply different preprocessing functions
    method_functions = { 
        'Base filters': apply_BF,
        'MWF': apply_MWF_pipeline, 
        'IClabel': apply_IC, 
        'wICA': apply_wICA, 
    }

    if method not in method_functions:
        raise ValueError(f"unknown method: {method}")

    return method_functions[method](data)

# functions to communicate with the excelfile 
def load_processed_filenames(excel_path):
    # retrieves the names of the files that are already processed
    processed_files = set()
    if os.path.exists(excel_path):
        all_sheets = pd.read_excel(excel_path, sheet_name=None) #add specific method in sheet name if testing new method
        for df in all_sheets.values():
            if 'filename' in df.columns:
                filenames = df['filename'].dropna().tolist()
            else:
                filenames = df.index.dropna().tolist()
            processed_files.update(filenames)
    return processed_files

def append_new_metrics_to_excel(metrics_frame):
    # saves the quantitative metrics to the Excel file
    if not os.path.exists(results_file):
        with pd.ExcelWriter(results_file, engine = "openpyxl") as writer:
            for method, metrics_list in metrics_frame.items():
                if metrics_list:
                    df = pd.DataFrame(metrics_list)
                    df.to_excel(writer, sheet_name=method)
    else:
        book = load_workbook(results_file)
        with pd.ExcelWriter(results_file, engine = "openpyxl", mode='a', if_sheet_exists="overlay") as writer:
            for method, metrics_list in metrics_frame.items():
                if metrics_list:
                    df = pd.DataFrame(metrics_list)
                    if method in book.sheetnames:
                        sheet = book[method]
                        startrow = sheet.max_row
                    else:
                        startrow = 0
                    df.to_excel(writer, sheet_name=method, header=startrow==0, startrow=startrow)

# functions visualisation
def get_individual_alpha_band(freqs, psd, fmin=6, fmax=12, width=4):
    # determine individual alpha frequency band 
    picks = mne.pick_channels(psd.info['ch_names'], include=['E70', 'E75', 'E83'])
    psd_selected = psd.data[picks, :]
    psd_mean = psd_selected.mean(axis=0)
    mask = (freqs >= fmin) & (freqs <= fmax)
    peak_freq = freqs[mask][np.argmax(psd_mean[mask])]
    return [peak_freq - width/2, peak_freq + width/2]

def band_power_from_psd(freqs, psd, band):
    # calculate mean power from given frequency band
    band_mask = (freqs >= band[0]) & (freqs <= band[1])
    return psd[:, band_mask].mean(axis=1)

def normalize_psd(psd_data, freqs, fmin=1, fmax=40):
    # normalize PSD by dividing by total power in specified range
    mask = (freqs >= fmin) & (freqs <= fmax)
    total_power = np.sum(psd_data[:, mask], axis=1, keepdims=True)
    return psd_data / total_power
#%% 
# Main analysis
eeg_files = os.listdir(EEG_folder)
eeg_files = [x for x in eeg_files if x.endswith("_convert.cdt.dpa")]

processed_files = load_processed_filenames(results_file)
counter = 0

for eeg_file in eeg_files:
    filenamebase = eeg_file.split('_')[0] + '_' + eeg_file.split('_')[3] #adjust if the EEG file names are build differently
    print(f'Evaluating based on file: {filenamebase}')
    
    if filenamebase in processed_files:
        print('File is already evaluated, continue with next file')
        continue

    # load original EEG files and crop resting state
    raw, event_raw, event_dict, markers = load_data_with_markers(eeg_file)
    raw.load_data()

    if markers==False:
        print("geen markers")
        continue 

    open_periods, closed_periods = identify_periods(raw)
    if not open_periods or not closed_periods:
        print('continue with next measurement')
        continue 
    start_RS = min(min(start for start, _ in closed_periods), min(start for start, _ in open_periods))
    end_RS = max(max(stop for _, stop in open_periods), max(stop for _, stop in closed_periods))
    RS_orig = raw.copy().crop(tmin=start_RS, tmax=end_RS)
    RS_orig.load_data()

    # Evaluate EEG files with some base filters for comparison
    RS_filt = apply_BF(RS_orig)
    quality_filt = assess_eeg_quality(RS_filt.copy().filter(l_freq=1, h_freq=100, verbose='CRITICAL'))
    metrics = {
        'quality_ratio': quality_filt['quality_ratio'],
        'filename': filenamebase
    }
    new_metrics_buffer['base_filters'].append(metrics)

    data_open = extract_segments(RS_filt, open_periods)
    data_closed = extract_segments(RS_filt, closed_periods)
    raw_open = mne.io.RawArray(data_open, RS_filt.info)
    raw_closed = mne.io.RawArray(data_closed, RS_filt.info)
    psd_closed = raw_closed.compute_psd(fmin=2, fmax=15, picks='eeg', n_fft = 1024)
    psd_closed.save(folder_PSDS + "/" + "base filters" + "/" + filenamebase + "_closed.h5", overwrite=True)
    psd_open = raw_open.compute_psd(fmin=2, fmax=15, picks='eeg', n_fft = 1024)
    psd_open.save(folder_PSDS + "/" + "base filters" + "/" + filenamebase + "_open.h5", overwrite=True)

    # Evaluate methods based on defined functions
    for method in pipelines:
        RS = apply_method(RS_orig, method)
   
        quality = assess_eeg_quality(RS.copy().filter(l_freq=1, h_freq=100, verbose='CRITICAL'))
        SER, ARR, cor, MI = compute_ser_arr(RS_filt, RS, quality_filt['bad_segments'])
        metrics = {
            'filename': filenamebase
            'quality_ratio': quality['quality_ratio'],
            'SER': SER,
            'ARR': ARR,
            'cor': cor,
            'MI': MI, 
        }
        new_metrics_buffer[method].append(metrics)
        
        data_open = extract_segments(RS, open_periods)
        data_closed = extract_segments(RS, closed_periods)
        raw_open = mne.io.RawArray(data_open, RS.info)
        raw_closed = mne.io.RawArray(data_closed, RS.info)
        psd_closed = raw_closed.compute_psd(fmin=2, fmax=15, picks='eeg', n_fft = 1024)
        psd_closed.save(folder_PSDS + "/" + method + "/" + filenamebase + "_closed.h5", overwrite=True)
        psd_open = raw_open.compute_psd(fmin=2, fmax=15, picks='eeg', n_fft = 1024)
        psd_open.save(folder_PSDS + "/" + method + "/" + filenamebase + "_open.h5", overwrite=True)

    # Evaluate methods based on preprocessed EEG's (copy this part to evaluate more pipelines in this way (also name them on the three other places above and below))
    raw_prep = mne.io.read_raw_fif(EEG_pipeline_results + "/" + filenamebase + "_prepped.fif", preload=True) # adjust if the eeg filenames are built differently 
    open_periods, closed_periods = identify_periods(raw_prep)

    if not open_periods or not closed_periods:
        print('continue with next measurement')
        continue 

    start_RS = min(min(start for start, _ in closed_periods), min(start for start, _ in open_periods))
    end_RS = max(max(stop for _, stop in open_periods), max(stop for _, stop in closed_periods))
    RS_prep = raw_prep.copy().crop(tmin=start_RS, tmax=end_RS)

    quality_prep = assess_eeg_quality(RS_prep.copy().filter(l_freq=1, h_freq=100, verbose='CRITICAL'))
    SER, ARR, cor, MI = compute_ser_arr(RS_filt, RS_prep, quality_filt['bad_segments'])
    metrics = {
        'filename': filenamebase
        'quality_ratio': quality_prep['quality_ratio'],
        'SER': SER,
        'ARR': ARR,
        'cor': cor,
        'MI': MI, 
    }
    new_metrics_buffer['preprocessing_pipeline'].append(metrics)

    data_open = extract_segments(RS_prep, open_periods)
    data_closed = extract_segments(RS_prep, closed_periods)
    raw_open = mne.io.RawArray(data_open, RS_prep.info)
    raw_closed = mne.io.RawArray(data_closed, RS_prep.info)
    psd_closed = raw_closed.compute_psd(fmin=2, fmax=15, picks='eeg', n_fft = 1024)
    psd_closed.save(folder_PSDS + "/" + "preprocessing_pipeline" + "/" + filenamebase + "_closed.h5", overwrite=True)
    psd_open = raw_open.compute_psd(fmin=2, fmax=15, picks='eeg', n_fft = 1024)
    psd_open.save(folder_PSDS + "/" + "preprocessing_pipeline" + "/" + filenamebase + "_open.h5", overwrite=True)

    # Save results
    counter += 1
    if counter % save_interval == 0:
        append_new_metrics_to_excel(new_metrics_buffer)
        new_metrics_buffer = {method: [] for method in pipelines}
        new_metrics_buffer['base_filters'] = []
        new_metrics_buffer['preprocessing_pipeline'] = [] #add pipelines like this if analysing based on preprocessed files instead of giving the function (also three more places above)

if any(len(metrics_list) > 0 for metrics_list in new_metrics_buffer.values()):
    append_new_metrics_to_excel(new_metrics_buffer)

print("Quality metrics updated! all files are evaluated, continue with visualisation of the results")  

#%%
# visualisation of cleaning performance 
%matplotlib Tk
all_sheets = pd.read_excel(excel_path, sheet_name = None)
methods = [m for m in all_sheets.keys()]

## 1. scatter plots of signal preservation against artifact removal for all pipelines together

data = []
for method, df in all_sheets.items():
    if method in ['MWF', 'IClabel', 'wICA', 'preprocessing_pipeline']:
        df = df.copy()
        df['method'] = method
        df['file'] = df.index
        data.append(df)
data = pd.concat(data, ignore_index=True)

fig, axes = plt.subplots(1,2, figsize=(10,5))

sns.scatterplot(data=data, x='SER', y='ARR', hue='method', palette='tab10', s=60, edgecolor='black', ax=axes[0])
axes[0].set_xlabel(f'Signal to Error Ratio (SER)')
axes[0].set_ylabel(f'Artifact to Residue Ratio (ARR)')
axes[0].grid(True)

sns.scatterplot(data=data, x='cor', y='MI', hue='method', palette='tab10', s=60, edgecolor='black', ax=axes[1])
plt.xlabel(f'Correlation before and after preprocessing')
plt.ylabel(f'Mutual information removed and preprocessed signal')
axes[1].grid(True)

handles, labels = axes[0].get_legend_handles_labels()
fig.legend(handles, labels, title='Method', bbox_to_anchor=(1.05, 1), loc='upper left')

plt.show()

## 2. paired scatter plots for improvements in data quality per pipeline

basis_df = all_sheets['base_filters'][['quality_ratio']].copy()
basis_df.columns = ['quality_ratio_baseline']

for method in [m for m in methods if m != 'base_filters']: 
    method_df = all_sheets[method][['quality_ratio']].copy()
    method_df.columns = ['quality_ratio_method']
    
    combined = pd.concat([basis_df, method_df], axis=1).dropna()
    difference = combined['quality_ratio_method'] - combined['quality_ratio_baseline']

    shapiro_stat, shapiro_p = shapiro(difference)
    if shapiro_p > 0.05:
        stat, p_value = ttest_rel(combined['quality_ratio_method'], combined['quality_ratio_baseline'])
        test_type = 'paired t-test'

        mean_diff = np.mean(difference)
        std_diff = np.std(difference, ddof=1)
        effect_size = mean_diff / std_diff
        effect_label = f"Cohen's d = {effect_size:.2f}"
    else: 
        stat, p_value = wilcoxon(combined['quality_ratio_method'], combined['quality_ratio_baseline'])
        test_type = 'Wilcoxon signed-rank test'
        n_pos = sum(difference > 0)
        n_neg = sum(difference < 0)
        effect_size = (n_pos - n_neg) / (n_pos + n_neg) if (n_pos + n_neg) != 0 else np.nan
        median_diff = np.median(difference)
        effect_label = f"rank biserial = {effect_size:.2f}, median diff = {median_diff:.2f}"

    plt.figure(figsize=(8,6))
    sns.stripplot(data=combined.melt(var_name='Moment', value_name='Quality Ratio'), 
                    x='Moment', y='Quality Ratio', jitter=True, alpha=0.7, size=6, palette='pastel')

    for i in range(len(combined)):
        plt.plot(['quality_ratio_baseline', 'quality_ratio_method'], combined.iloc[i].values, color='gray', alpha=0.4)

    plt.title(f'{method} - Quality ratio\n{test_type} p={p_value:.3f}\n{effect_label}')
    plt.ylabel('Quality ratio')
    plt.xticks([0,1], ['current CBL preprocessing', f'{method} preprocesing pipeline'])
    plt.tight_layout()
    plt.grid(True)
    plt.show()

## 3. topomaps and spectra for detection of berger effect with bayesian t-test to address the strength
results = {}
bayes_stats = {}

for method in methods:
    method_path = os.path.join(folder_PSDS, method)
    files = sorted([f for f in os.listdir(method_path) if f.endswith('_open.h5')])
    results[method] = {'E75_open': [], 'E75_closed': [], 'alpha_open': [], 'alpha_closed': [], 'rms_open': [], 'rms_closed': []}

    for f_open in files:
        f_closed = f_open.replace('_open.h5', '_closed.h5')
        path_open = os.path.join(method_path, f_open)
        path_closed = os.path.join(method_path, f_closed)
        if not os.path.exists(path_closed):
            continue

        psd_open = read_spectrum(path_open)
        psd_array_open = psd_open.get_data()
        freqs = psd_open.freqs
        ch_names = psd_open.info.ch_names
        psd_closed = read_spectrum(path_closed)
        psd_array_closed = psd_closed.get_data()
        psd_open_normalized = normalize_psd(psd_array_open, freqs)
        psd_closed_normalized = normalize_psd(psd_array_closed, freqs)

        idx_E75 = ch_names.index('E75')
        results[method]['E75_open'].append(psd_open_normalized[idx_E75])
        results[method]['E75_closed'].append(psd_closed_normalized[idx_E75])

        ind_alpha_band = get_individual_alpha_band(freqs, psd_closed)
        alpha_open = band_power_from_psd(freqs, psd_open_normalized, ind_alpha_band)
        alpha_closed = band_power_from_psd(freqs, psd_closed_normalized, ind_alpha_band)
        results[method]['alpha_open'].append(alpha_open)
        results[method]['alpha_closed'].append(alpha_closed)

        results[method]['rms_open'].append(np.sqrt(np.mean(alpha_open**2)))
        results[method]['rms_closed'].append(np.sqrt(np.mean(alpha_closed**2)))

    # Plot mean psd of E75 over all measurements for pipeline
    mean_open = np.mean(results[method]['E75_open'], axis=0)
    std_open = np.std(results[method]['E75_open'], axis=0)
    mean_closed = np.mean(results[method]['E75_closed'], axis=0)
    std_closed = np.std(results[method]['E75_closed'], axis=0)

    plt.figure(figsize=(10, 5))
    plt.title(f'PSD E75 - {method}')
    plt.plot(freqs, mean_open, label='Open', color='blue')
    plt.fill_between(freqs, mean_open - std_open, mean_open + std_open, color='blue', alpha=0.3)
    plt.plot(freqs, mean_closed, label='Closed', color='orange')
    plt.fill_between(freqs, mean_closed - std_closed, mean_closed + std_closed, color='orange', alpha=0.3)
    plt.xlabel('Frequency (Hz)')
    plt.ylabel('Power (normalized)')
    plt.legend()
    plt.tight_layout()
    plt.show()

    # Topomap alpha power
    if method == 'base_filters':
        kept_channels = [ch for ch in ch_names if ch != 'E55']
        info = mne.create_info(ch_names=kept_channels, sfreq=250, ch_types='eeg')
    else: 
        info = mne.create_info(ch_names=ch_names, sfreq=250, ch_types='eeg')
    montage = mne.channels.make_standard_montage('GSN-HydroCel-128')
    info.set_montage(montage)

    mean_topo_open = np.mean(results[method]['alpha_open'], axis=0)
    mean_topo_closed = np.mean(results[method]['alpha_closed'], axis=0)

    alpha_power_prep = np.concatenate([mean_topo_open, mean_topo_closed])
    vlim_prep = [alpha_power_prep.min(), alpha_power_prep.max()]

    fig, ax = plt.subplots(1, 2, figsize=(10, 4))
    im1, _ = plot_topomap(mean_topo_open, info, axes=ax[0], vlim = vlim_prep, show=False)
    ax[0].set_title('Open')
    im2, _ = plot_topomap(mean_topo_closed, info, axes=ax[1], vlim = vlim_prep, show=False)
    ax[1].set_title('Closed')
    plt.suptitle(f'Mean alpha power topomap - {method}')
    cbar = fig.colorbar(im2, ax=ax)
    cbar.set_label('Power (normalized)')
    plt.show

    # Bayesian paired t-test
    open_rms = results[method]['rms_open']
    closed_rms = results[method]['rms_closed']
    result = pg.ttest(closed_rms, open_rms, paired=True)
    bayes_stats[method] = {
        'BF_10': result['BF10'],
        'cohens_d': result['cohen-d'],
        'p': result['p-val']
    }

# Print Bayesian t test results
print("\Results Bayesian t-test per method:")
for method, stats in bayes_stats.items():
    print(f"\n{method}:")
    print(f"BF10: {stats['BF_10']}")
    print(f"cohen's d: {stats['cohens_d']}")
    print(f"p-value: {stats['p']}")
